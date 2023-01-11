from openpyxl import load_workbook
import datetime
import time

# Load the Excel file
wb = load_workbook('translations.xlsx')
sheet = wb.active

# Get the current date
today = datetime.date.today()
day = today.day
# time = datetime.now().strftime("%H:%M")
time_now = time.strftime("%H:%M")


# Get the word, translation, and pronunciation for the current day
word = sheet.cell(row=day, column=1).value
translation = sheet.cell(row=day, column=2).value
pronunciation = sheet.cell(row=day, column=3).value
